"""
Excel File Manipulation with Python
This module provides comprehensive functionality for manipulating Excel files using Python.
It includes features for reading, writing, modifying Excel files and SQL integration.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import pandas as pd
from sqlalchemy import create_engine
import os


class ExcelManipulator:
    """A class to handle various Excel file manipulation operations."""
    
    def __init__(self, filename):
        """
        Initialize the ExcelManipulator with a filename.
        
        Args:
            filename (str): The name of the Excel file to work with
        """
        self.filename = filename
        self.workbook = None
        self.worksheet = None
    
    def create_workbook(self):
        """Create a new Excel workbook."""
        self.workbook = openpyxl.Workbook()
        self.worksheet = self.workbook.active
        self.worksheet.title = "Sheet1"
        print(f"New workbook created: {self.filename}")
    
    def load_workbook(self):
        """Load an existing Excel workbook."""
        if not os.path.exists(self.filename):
            raise FileNotFoundError(f"File {self.filename} not found")
        self.workbook = openpyxl.load_workbook(self.filename)
        self.worksheet = self.workbook.active
        print(f"Workbook loaded: {self.filename}")
    
    def save_workbook(self):
        """Save the current workbook."""
        if self.workbook:
            self.workbook.save(self.filename)
            print(f"Workbook saved: {self.filename}")
        else:
            print("No workbook to save")
    
    def write_data(self, row, col, value):
        """
        Write data to a specific cell.
        
        Args:
            row (int): Row number (1-indexed)
            col (int): Column number (1-indexed)
            value: Value to write to the cell
        """
        if self.worksheet:
            self.worksheet.cell(row=row, column=col, value=value)
            print(f"Data written to cell ({row}, {col}): {value}")
        else:
            print("No worksheet available")
    
    def read_data(self, row, col):
        """
        Read data from a specific cell.
        
        Args:
            row (int): Row number (1-indexed)
            col (int): Column number (1-indexed)
            
        Returns:
            The value in the specified cell
        """
        if self.worksheet:
            value = self.worksheet.cell(row=row, column=col).value
            print(f"Data read from cell ({row}, {col}): {value}")
            return value
        else:
            print("No worksheet available")
            return None
    
    def write_row(self, row_num, data_list):
        """
        Write a list of values to a row.
        
        Args:
            row_num (int): Row number (1-indexed)
            data_list (list): List of values to write
        """
        if self.worksheet:
            for col_num, value in enumerate(data_list, start=1):
                self.worksheet.cell(row=row_num, column=col_num, value=value)
            print(f"Row {row_num} written with {len(data_list)} values")
        else:
            print("No worksheet available")
    
    def write_column(self, col_num, data_list):
        """
        Write a list of values to a column.
        
        Args:
            col_num (int): Column number (1-indexed)
            data_list (list): List of values to write
        """
        if self.worksheet:
            for row_num, value in enumerate(data_list, start=1):
                self.worksheet.cell(row=row_num, column=col_num, value=value)
            print(f"Column {col_num} written with {len(data_list)} values")
        else:
            print("No worksheet available")
    
    def format_cell(self, row, col, font_bold=False, font_size=11, bg_color=None):
        """
        Apply formatting to a cell.
        
        Args:
            row (int): Row number (1-indexed)
            col (int): Column number (1-indexed)
            font_bold (bool): Whether to make the font bold
            font_size (int): Font size
            bg_color (str): Background color in hex format (e.g., "FFFF00" for yellow)
        """
        if self.worksheet:
            cell = self.worksheet.cell(row=row, column=col)
            cell.font = Font(bold=font_bold, size=font_size)
            if bg_color:
                cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')
            print(f"Cell ({row}, {col}) formatted")
        else:
            print("No worksheet available")
    
    def read_all_data(self):
        """
        Read all data from the worksheet.
        
        Returns:
            list: List of lists containing all cell values
        """
        if self.worksheet:
            data = []
            for row in self.worksheet.iter_rows(values_only=True):
                data.append(list(row))
            print(f"Read {len(data)} rows from worksheet")
            return data
        else:
            print("No worksheet available")
            return []
    
    def add_formula(self, row, col, formula):
        """
        Add a formula to a cell.
        
        Args:
            row (int): Row number (1-indexed)
            col (int): Column number (1-indexed)
            formula (str): Excel formula (e.g., "=SUM(A1:A10)")
        """
        if self.worksheet:
            self.worksheet.cell(row=row, column=col, value=formula)
            print(f"Formula added to cell ({row}, {col}): {formula}")
        else:
            print("No worksheet available")


class ExcelSQLIntegration:
    """A class to handle Excel and SQL database integration."""
    
    @staticmethod
    def excel_to_dataframe(filename, sheet_name=0):
        """
        Convert Excel file to pandas DataFrame.
        
        Args:
            filename (str): Path to Excel file
            sheet_name (str or int): Sheet name or index
            
        Returns:
            pandas.DataFrame: DataFrame containing the Excel data
        """
        df = pd.read_excel(filename, sheet_name=sheet_name)
        print(f"Excel file loaded into DataFrame: {filename}")
        print(f"Shape: {df.shape}")
        return df
    
    @staticmethod
    def dataframe_to_excel(df, filename, sheet_name='Sheet1', index=False):
        """
        Save pandas DataFrame to Excel file.
        
        Args:
            df (pandas.DataFrame): DataFrame to save
            filename (str): Output Excel file path
            sheet_name (str): Sheet name
            index (bool): Whether to write row indices
        """
        df.to_excel(filename, sheet_name=sheet_name, index=index)
        print(f"DataFrame saved to Excel: {filename}")
    
    @staticmethod
    def excel_to_sql(excel_file, table_name, db_path='database.db', sheet_name=0):
        """
        Import Excel data into SQL database.
        
        Args:
            excel_file (str): Path to Excel file
            table_name (str): Name of the SQL table
            db_path (str): Path to SQLite database
            sheet_name (str or int): Sheet name or index
        """
        # Read Excel file
        df = pd.read_excel(excel_file, sheet_name=sheet_name)
        
        # Create SQLite engine
        engine = create_engine(f'sqlite:///{db_path}')
        
        # Write DataFrame to SQL
        df.to_sql(table_name, engine, if_exists='replace', index=False)
        print(f"Data from {excel_file} imported to SQL table '{table_name}' in {db_path}")
    
    @staticmethod
    def sql_to_excel(query, excel_file, db_path='database.db', sheet_name='Sheet1'):
        """
        Export SQL query results to Excel file.
        
        Args:
            query (str): SQL query to execute
            excel_file (str): Output Excel file path
            db_path (str): Path to SQLite database
            sheet_name (str): Sheet name
        """
        # Create SQLite engine
        engine = create_engine(f'sqlite:///{db_path}')
        
        # Execute query and load to DataFrame
        df = pd.read_sql_query(query, engine)
        
        # Save to Excel
        df.to_excel(excel_file, sheet_name=sheet_name, index=False)
        print(f"SQL query results exported to {excel_file}")


def example_usage():
    """Demonstrate various Excel manipulation operations."""
    
    print("=" * 60)
    print("Excel File Manipulation Examples")
    print("=" * 60)
    
    # Example 1: Create a new Excel file and write data
    print("\n1. Creating a new Excel file...")
    excel = ExcelManipulator("example_output.xlsx")
    excel.create_workbook()
    
    # Write headers
    headers = ["Name", "Age", "City", "Salary"]
    excel.write_row(1, headers)
    excel.format_cell(1, 1, font_bold=True, font_size=12, bg_color="CCCCCC")
    excel.format_cell(1, 2, font_bold=True, font_size=12, bg_color="CCCCCC")
    excel.format_cell(1, 3, font_bold=True, font_size=12, bg_color="CCCCCC")
    excel.format_cell(1, 4, font_bold=True, font_size=12, bg_color="CCCCCC")
    
    # Write data rows
    data = [
        ["John Doe", 30, "New York", 75000],
        ["Jane Smith", 25, "Los Angeles", 65000],
        ["Bob Johnson", 35, "Chicago", 80000],
        ["Alice Brown", 28, "Houston", 70000]
    ]
    
    for idx, row_data in enumerate(data, start=2):
        excel.write_row(idx, row_data)
    
    # Add a formula to calculate average salary
    excel.write_data(6, 3, "Average Salary:")
    excel.format_cell(6, 3, font_bold=True)
    excel.add_formula(6, 4, "=AVERAGE(D2:D5)")
    
    excel.save_workbook()
    
    # Example 2: Read data from Excel
    print("\n2. Reading data from Excel file...")
    excel2 = ExcelManipulator("example_output.xlsx")
    excel2.load_workbook()
    all_data = excel2.read_all_data()
    print(f"First few rows: {all_data[:3]}")
    
    # Example 3: Using pandas for Excel manipulation
    print("\n3. Using pandas for advanced operations...")
    df = ExcelSQLIntegration.excel_to_dataframe("example_output.xlsx")
    print("\nDataFrame content:")
    print(df.head())
    
    # Perform some data analysis
    print(f"\nAverage salary: ${df['Salary'].mean():.2f}")
    print(f"Max salary: ${df['Salary'].max():.2f}")
    print(f"Min salary: ${df['Salary'].min():.2f}")
    
    # Example 4: Excel to SQL and back
    print("\n4. Excel to SQL integration...")
    ExcelSQLIntegration.excel_to_sql("example_output.xlsx", "employees", "company.db")
    
    # Query data from SQL and export to new Excel file
    query = "SELECT * FROM employees WHERE Salary > 70000"
    ExcelSQLIntegration.sql_to_excel(query, "high_salary_employees.xlsx", "company.db")
    
    print("\n" + "=" * 60)
    print("Examples completed successfully!")
    print("=" * 60)


if __name__ == "__main__":
    example_usage()
